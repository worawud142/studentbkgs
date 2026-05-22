#!/usr/bin/env python3
import json
import re
import sys
from datetime import datetime, date
from pathlib import Path

from openpyxl import load_workbook


ALIASES = {
    "studentCode": {
        "studentcode",
        "studentid",
        "studentno",
        "code",
        "รหัส",
        "รหัสนักเรียน",
        "รหัสประจำตัวนักเรียน",
    },
    "prefix": {
        "prefix",
        "คำนำหน้า",
        "คำนำหน้าชื่อ",
    },
    "firstName": {
        "firstname",
        "firstชื่อ",
        "ชื่อ",
        "ชื่อจริง",
    },
    "lastName": {
        "lastname",
        "นามสกุล",
        "surname",
        "ชื่อสกุล",
    },
    "fullName": {
        "fullname",
        "ชื่อ-สกุล",
        "ชื่อสกุล",
        "ชื่อและนามสกุล",
        "name",
        "นักเรียน",
    },
    "nationalId": {
        "nationalid",
        "idcard",
        "เลขประชาชน",
        "เลขบัตรประชาชน",
        "เลขประจำตัวประชาชน",
    },
    "birthDate": {
        "birthdate",
        "dob",
        "วันเกิด",
        "วัน/เดือน/ปีเกิด",
    },
    "gender": {
        "gender",
        "เพศ",
    },
    "studentNumber": {
        "studentnumber",
        "no",
        "เลขที่",
        "ลำดับ",
        "number",
    },
    "status": {
        "status",
        "สถานะ",
    },
}

PREFIXES = ("เด็กชาย", "เด็กหญิง", "นาย", "นางสาว", "ด.ช.", "ด.ญ.")


def normalize(value):
    if value is None:
      return ""
    text = str(value).strip().lower()
    return re.sub(r"[\s\-_\/\.\(\)\[\]{}:]+", "", text)


def clean_text(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def format_date(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = clean_text(value)
    if not text:
        return ""
    if re.match(r"^\d{4}-\d{2}-\d{2}$", text):
        return text
    return text


def format_number(value):
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    text = clean_text(value)
    if not text:
        return None
    try:
        if "." in text:
            num = float(text)
            return int(num) if num.is_integer() else None
        return int(text)
    except Exception:
        return None


def map_gender(value):
    text = normalize(value)
    if text in {"m", "male", "ชาย", "ช"}:
        return "male"
    if text in {"f", "female", "หญิง", "ญ"}:
        return "female"
    return ""


def map_status(value):
    text = normalize(value)
    if text in {"active", "ใช้งาน", "ปกติ"}:
        return "active"
    if text in {"transferred", "ย้าย", "ย้ายออก"}:
        return "transferred"
    if text in {"graduated", "จบ", "จบการศึกษา"}:
        return "graduated"
    if text in {"dropped", "drop", "ลาออก"}:
        return "dropped"
    return ""


def split_full_name(full_name):
    text = clean_text(full_name)
    if not text:
        return "", "", ""

    prefix = ""
    remaining = text
    for candidate in PREFIXES:
        if remaining.startswith(candidate + " "):
            prefix = candidate
            remaining = remaining[len(candidate):].strip()
            break

    parts = [part for part in re.split(r"\s+", remaining) if part]
    if len(parts) >= 2:
        first_name = parts[0]
        last_name = " ".join(parts[1:])
    elif len(parts) == 1:
        first_name = parts[0]
        last_name = ""
    else:
        first_name = ""
        last_name = ""

    return prefix, first_name, last_name


def find_header_row(rows):
    best_index = None
    best_map = {}

    for index, row in enumerate(rows):
        mapping = {}
        non_empty = 0
        for col_index, value in enumerate(row):
            if value is None or str(value).strip() == "":
                continue
            non_empty += 1
            normalized = normalize(value)
            for canonical, aliases in ALIASES.items():
                if normalized in {normalize(alias) for alias in aliases}:
                    mapping[canonical] = col_index
                    break

        if "studentCode" in mapping and ("firstName" in mapping or "fullName" in mapping):
            return index, mapping
        if best_index is None and mapping and non_empty >= 2:
            best_index = index
            best_map = mapping

    return best_index, best_map


def main():
    if len(sys.argv) < 2:
        print(json.dumps({"error": "usage: excel_importer.py <workbook-path>"}))
        return 1

    workbook_path = Path(sys.argv[1])
    wb = load_workbook(workbook_path, data_only=True, read_only=True)

    selected_sheet = None
    header_row = None
    header_map = {}
    sheet_rows = None

    for sheet in wb.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        header_row, header_map = find_header_row(rows[:20])
        if header_row is not None and header_map:
            selected_sheet = sheet.title
            sheet_rows = rows
            break

    if selected_sheet is None or sheet_rows is None:
        print(json.dumps({"sheetName": None, "rows": [], "errors": [{"rowNumber": 0, "message": "ไม่พบแถวหัวตารางในไฟล์ Excel"}]}))
        return 0

    parsed_rows = []
    errors = []
    seen_codes = set()

    for row_index in range(header_row + 1, len(sheet_rows)):
        row = sheet_rows[row_index]
        if not row or all(value is None or str(value).strip() == "" for value in row):
            continue

        record = {
            "rowNumber": row_index + 1,
            "studentCode": "",
            "prefix": "",
            "firstName": "",
            "lastName": "",
            "nationalId": "",
            "birthDate": "",
            "gender": "",
            "studentNumber": None,
            "status": "active",
        }

        for canonical, col_index in header_map.items():
            if col_index >= len(row):
                continue
            value = row[col_index]
            if canonical == "studentNumber":
                record[canonical] = format_number(value)
            elif canonical == "birthDate":
                record[canonical] = format_date(value)
            elif canonical == "gender":
                record[canonical] = map_gender(value)
            elif canonical == "status":
                record[canonical] = map_status(value) or "active"
            else:
                record[canonical] = clean_text(value)

        if not record["firstName"] and not record["lastName"]:
            prefix, first_name, last_name = split_full_name(record.get("fullName") or "")
            if prefix and not record["prefix"]:
                record["prefix"] = prefix
            if first_name and not record["firstName"]:
                record["firstName"] = first_name
            if last_name and not record["lastName"]:
                record["lastName"] = last_name

        if not record["studentCode"]:
            errors.append({"rowNumber": record["rowNumber"], "message": "ไม่พบรหัสนักเรียน"})
            continue
        if record["studentCode"] in seen_codes:
            errors.append({"rowNumber": record["rowNumber"], "message": "รหัสนักเรียนซ้ำในไฟล์"})
            continue
        if not record["firstName"] or not record["lastName"]:
            errors.append({"rowNumber": record["rowNumber"], "message": "ต้องมีชื่อและนามสกุล"})
            continue

        seen_codes.add(record["studentCode"])
        parsed_rows.append(record)

    print(json.dumps({
        "sheetName": selected_sheet,
        "rows": parsed_rows,
        "errors": errors,
    }, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    sys.exit(main())
