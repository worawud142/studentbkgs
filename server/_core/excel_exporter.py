#!/usr/bin/env python3
import json
import os
import re
import sys
from copy import copy
from datetime import date, timedelta

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter


def cell_to_row_col(cell_ref):
    col = ""
    row = ""
    for ch in cell_ref:
        if ch.isalpha():
            col += ch
        elif ch.isdigit():
            row += ch
    return int(row), col


def clear_rect(ws, min_row, max_row, min_col, max_col):
    for row in ws.iter_rows(
        min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
    ):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            cell.value = None


def write_table(ws, start_row, headers, rows):
    for index, header in enumerate(headers, start=1):
        ws.cell(row=start_row, column=index, value=header)
    for row_index, row in enumerate(rows, start=start_row + 1):
        for col_index, header in enumerate(headers, start=1):
            ws.cell(row=row_index, column=col_index, value=row.get(header, ""))


def ensure_sheet(wb, title):
    if title in wb.sheetnames:
        return wb[title]
    return wb.create_sheet(title)


def clone_worksheet(wb, source_name, new_name):
    if new_name in wb.sheetnames:
        del wb[new_name]
    source = wb[source_name]
    target = wb.copy_worksheet(source)
    target.title = new_name
    return target


def normalize_text(value):
    return re.sub(r"[\s\-_(){}\[\]\/\\.,:;]+", "", str(value or "").lower())


def normalize_cell_value(value):
    if value is None:
        return ""
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value
    return str(value)


def excel_number_value(value):
    if value is None or value == "":
        return ""
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return value
    if numeric.is_integer():
        return int(numeric)
    return numeric


def write_excel_number(cell, value):
    normalized = excel_number_value(value)
    cell.value = normalized
    if normalized == "":
        return
    if isinstance(normalized, int):
        cell.number_format = "0"
    elif isinstance(normalized, float):
        cell.number_format = "0.##"


def sheet_search_text(ws):
    values = []
    cell = ws["A3"]
    if cell.value is not None and not (isinstance(cell.value, str) and cell.value.startswith("=")):
        values.append(str(cell.value))
    return normalize_text(" ".join(values))


def find_subject_sheet(wb, assignment):
    aliases = [
        assignment.get("subjectName", ""),
        assignment.get("subjectCode", ""),
        assignment.get("subjectGroup", ""),
    ]
    aliases = [normalize_text(alias) for alias in aliases if normalize_text(alias)]
    common_sheets = {
        "ปก",
        "ข้อมูลนร.2-3",
        "เวลาเรียน4-16",
        "ส่งออกข้อมูลจริง",
        "เกรดรวม59-60",
        "คำอธิบายการใช้ 61",
        "คำอธิบายการใช้ 62",
        "ผลการเรียน",
        "ผลการเรียนรู้",
    }

    for ws in wb.worksheets:
        if ws.title in common_sheets or ws.title.startswith("คำอธิบาย"):
            continue
        haystack = normalize_text(f"{ws.title} {sheet_search_text(ws)}")
        if any(alias in haystack for alias in aliases):
            return ws
    return None


def is_primary_template(template_name, assignment=None):
    assignment = assignment or {}
    return (
        template_name == "เก็บคะแนนประถม.xlsx"
        or template_name.startswith("ปพ.5-ป")
        or assignment.get("classroomLevel") == "primary"
    )


def is_secondary_template(template_name, assignment=None):
    assignment = assignment or {}
    return (
        template_name == "ตัวอย่างมัธยม.xlsx"
        or template_name.startswith("ปพ.5-ม")
        or assignment.get("classroomLevel") == "secondary"
    )


PRIMARY_UNIT_COLUMN_RANGES = {
    "midyear": (3, 12),
    "endyear": (18, 27),
}

PRIMARY_FINAL_COLUMNS = {
    "midyear": 14,   # N = ปลายภาค 1
    "endyear": 29,   # AC = ปลายภาค 2
}

LATEST_PRIMARY_TERM_SHEETS = {
    "midyear": "ภาค1(8)",
    "endyear": "ภาค2 (9)",
}

LATEST_PRIMARY_SUMMARY_SHEET = "สรุปผลรวม (10)"
LATEST_PRIMARY_UNIT_SCORE_COLUMNS = [3, 8, 13, 18, 23]
LATEST_PRIMARY_UNIT_SUMMARY_COLUMNS = {
    "midyear": [3, 4, 5, 6, 7],
    "endyear": [9, 10, 11, 12, 13],
}
LATEST_PRIMARY_FINAL_SUMMARY_COLUMNS = {
    "midyear": 15,
    "endyear": 16,
}

LATEST_SECONDARY_UNIT_SHEETS = [
    ("หน่วย 1,4 (5)", [3, 10, 17, 24]),
    ("หน่วย 5,8 (6)", [3, 10, 17, 24]),
    ("หน่วย 9,12 (7)", [3, 10, 17, 24]),
]
LATEST_SECONDARY_SUMMARY_SHEET = "สรุปผลรวม (8)"
LATEST_SECONDARY_UNIT_SUMMARY_COLUMNS = list(range(3, 15))
LATEST_SECONDARY_FINAL_SUMMARY_COLUMNS = {
    "midyear": 15,
    "endyear": 16,
}

THAI_MONTHS = {
    "ม.ค.": 1,
    "ก.พ.": 2,
    "มี.ค.": 3,
    "เม.ย.": 4,
    "พ.ค.": 5,
    "มิ.ย.": 6,
    "ก.ค.": 7,
    "ส.ค.": 8,
    "ก.ย.": 9,
    "ต.ค.": 10,
    "พ.ย.": 11,
    "ธ.ค.": 12,
}

ATTENDANCE_MARKS = {
    "present": "/",
    "absent": "ข",
    "late": "ส",
    "excused": "ล",
}


def get_primary_unit_columns(term=None):
    if term:
        ranges = [PRIMARY_UNIT_COLUMN_RANGES[term]]
    else:
        ranges = list(PRIMARY_UNIT_COLUMN_RANGES.values())
    columns = []
    for start_col, end_col in ranges:
        for col in range(start_col, end_col + 1):
            columns.append(col)
    return columns


def clear_columns(ws, min_row, max_row, columns):
    for col in columns:
        for row in range(min_row, max_row + 1):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None


def clear_primary_subject_sheet(ws):
    clear_columns(ws, 6, ws.max_row, get_primary_unit_columns())
    clear_columns(ws, 6, ws.max_row, list(PRIMARY_FINAL_COLUMNS.values()))


def normalize_primary_term(category):
    term = str(category.get("term", "") or "").strip().lower()
    if term in {"midyear", "endyear"}:
        return term
    return "midyear"


def is_primary_fixed_final_category(category):
    return str(category.get("name", "") or "").strip() in {"ปลายภาค 1", "ปลายภาค 2", "กลางภาค", "ปลายภาค"}


def sort_primary_categories(categories):
    return sorted(
        categories,
        key=lambda category: (
            1 if normalize_primary_term(category) == "endyear" else 0,
            1 if is_primary_fixed_final_category(category) else 0,
            int(category.get("order", 0) or 0),
            str(category.get("name", "") or ""),
        ),
    )


def get_primary_final_category(categories, term):
    expected_name = "ปลายภาค 1" if term == "midyear" else "ปลายภาค 2"
    for category in categories:
        if str(category.get("name", "") or "").strip() == expected_name:
            return category
    return None


def get_secondary_final_category(categories, term):
    expected_names = {"midyear": {"กลางภาค", "ปลายภาค 1"}, "endyear": {"ปลายภาค", "ปลายภาค 2"}}[term]
    for category in categories:
        if str(category.get("name", "") or "").strip() in expected_names:
            return category
    for category in categories:
        if normalize_primary_term(category) == term and is_primary_fixed_final_category(category):
            return category
    return None


def has_latest_primary_layout(wb):
    return all(
        name in wb.sheetnames
        for name in ["ปก (1)", "เวลาเรียน (2)", "ภาค1(8)", "ภาค2 (9)", LATEST_PRIMARY_SUMMARY_SHEET]
    )


def has_latest_secondary_layout(wb):
    return all(name in wb.sheetnames for name in ["ปก (1)", "เวลาเรียน (2)", LATEST_SECONDARY_SUMMARY_SHEET])


def classroom_grade(assignment):
    grade = assignment.get("classroomGrade")
    if grade not in (None, ""):
        return grade
    match = re.search(r"(\d+)", str(assignment.get("classroomName", "") or ""))
    return int(match.group(1)) if match else ""


def write_if_not_merged(ws, row, column, value):
    cell = ws.cell(row=row, column=column)
    if isinstance(cell, MergedCell):
        return
    cell.value = value


def assigned_teacher_label(assignment):
    return (
        "ครูประจำชั้น"
        if str(assignment.get("classroomLevel", "") or "").strip() == "primary"
        else "ครูที่ปรึกษา"
    )


def teacher_name(assignment):
    return str(assignment.get("teacherName") or "").strip()


def homeroom_teacher_names(assignment):
    names = assignment.get("homeroomTeacherNames")
    if isinstance(names, list):
        return [str(name).strip() for name in names if str(name).strip()]
    name = str(assignment.get("homeroomTeacherName") or "").strip()
    return [part.strip() for part in name.split("/") if part.strip()]


def set_black_font(cell):
    font = copy(cell.font)
    font.color = "FF000000"
    cell.font = font


def write_latest_cover(wb, assignment):
    if "ปก (1)" not in wb.sheetnames:
        return
    ws = wb["ปก (1)"]
    write_if_not_merged(ws, 9, 7, classroom_grade(assignment))
    academic_year = assignment.get("academicYear") or {}
    write_if_not_merged(ws, 9, 15, academic_year.get("year", ""))
    write_if_not_merged(ws, 10, 5, assignment.get("subjectName", ""))
    write_if_not_merged(ws, 10, 12, assignment.get("subjectCode", ""))
    write_if_not_merged(ws, 11, 5, assignment.get("hoursPerWeek", ""))
    write_if_not_merged(ws, 11, 14, assignment.get("subjectCredits", ""))
    write_if_not_merged(ws, 12, 3, "ครูผู้สอน")
    write_if_not_merged(ws, 12, 5, teacher_name(assignment))
    for cell_ref in ["C12", "E12"]:
        set_black_font(ws[cell_ref])

    label = assigned_teacher_label(assignment)
    positions = [(13, 3, 5), (14, 3, 5), (13, 10, 12), (14, 10, 12)]
    for index, name in enumerate(homeroom_teacher_names(assignment)[: len(positions)]):
        row, label_col, name_col = positions[index]
        write_if_not_merged(ws, row, label_col, label)
        write_if_not_merged(ws, row, name_col, name)
        set_black_font(ws.cell(row=row, column=label_col))
        set_black_font(ws.cell(row=row, column=name_col))


def secondary_attribute_formula(summary_row):
    return (
        f"=IF('{LATEST_SECONDARY_SUMMARY_SHEET}'!Q{summary_row}>70,3,"
        f"IF('{LATEST_SECONDARY_SUMMARY_SHEET}'!Q{summary_row}>59,2,"
        f"IF('{LATEST_SECONDARY_SUMMARY_SHEET}'!Q{summary_row}>49,1,"
        f"IF('{LATEST_SECONDARY_SUMMARY_SHEET}'!Q{summary_row}<50,0))))"
    )


def repair_latest_assessment_formulas(wb, visible_students):
    sheet_name = "คุณลักษณะ อ่าน สมรรถนะ (9)"
    if sheet_name not in wb.sheetnames or LATEST_SECONDARY_SUMMARY_SHEET not in wb.sheetnames:
        return

    ws = wb[sheet_name]
    for row_offset, _student in enumerate(visible_students):
        assessment_row = 5 + row_offset
        summary_row = 7 + row_offset
        write_if_not_merged(ws, assessment_row, 11, secondary_attribute_formula(summary_row))


def write_latest_student_lists(wb, visible_students):
    for ws in wb.worksheets:
        if not ws.title.startswith("เวลาเรียน"):
            continue
        clear_rect(ws, 6, ws.max_row, 1, 3)
        for row, student in enumerate(visible_students, start=6):
            write_if_not_merged(ws, row, 1, student.get("studentNumber") or row - 5)
            write_if_not_merged(ws, row, 2, student.get("studentCode") or "")
            write_if_not_merged(ws, row, 3, student_visible_name(student))


def write_latest_score_student_names(wb, visible_students):
    sheet_rows = []
    if has_latest_primary_layout(wb):
        sheet_rows.extend(
            [
                (LATEST_PRIMARY_TERM_SHEETS["midyear"], 7),
                (LATEST_PRIMARY_TERM_SHEETS["endyear"], 7),
                (LATEST_PRIMARY_SUMMARY_SHEET, 8),
            ]
        )
    if has_latest_secondary_layout(wb):
        sheet_rows.extend(
            [(sheet_name, 6) for sheet_name, _ in LATEST_SECONDARY_UNIT_SHEETS]
        )
        sheet_rows.append((LATEST_SECONDARY_SUMMARY_SHEET, 7))

    for sheet_name, start_row in sheet_rows:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        clear_rect(ws, start_row, ws.max_row, 1, 2)
        for row, student in enumerate(visible_students, start=start_row):
            write_if_not_merged(ws, row, 1, student.get("studentNumber") or row - start_row + 1)
            write_if_not_merged(ws, row, 2, student_visible_name(student))


def extend_row_formulas(ws, source_row, target_start_row, student_count, min_col=1, max_col=None):
    if student_count <= 0:
        return
    max_col = max_col or ws.max_column
    source_formulas = []
    for col in range(min_col, max_col + 1):
        cell = ws.cell(row=source_row, column=col)
        if isinstance(cell.value, str) and cell.value.startswith("="):
            source_formulas.append((col, cell.coordinate, cell.value))

    for row in range(target_start_row, target_start_row + student_count):
        for col, origin, formula in source_formulas:
            target = ws.cell(row=row, column=col)
            if isinstance(target, MergedCell):
                continue
            target.value = Translator(formula, origin=origin).translate_formula(target.coordinate)


def extend_latest_student_formulas(wb, visible_students):
    student_count = len(visible_students)
    for ws in wb.worksheets:
        if ws.title.startswith("เวลาเรียน"):
            extend_row_formulas(ws, 6, 6, student_count)

    formula_sheets = []
    if has_latest_primary_layout(wb):
        formula_sheets.extend(
            [
                (LATEST_PRIMARY_TERM_SHEETS["midyear"], 7, 7),
                (LATEST_PRIMARY_TERM_SHEETS["endyear"], 7, 7),
                (LATEST_PRIMARY_SUMMARY_SHEET, 8, 8),
                ("คุณลักษณะ -อ่าน -สมรรถนะ(11)", 5, 5),
            ]
        )
    if has_latest_secondary_layout(wb):
        formula_sheets.extend(
            [(sheet_name, 6, 6) for sheet_name, _columns in LATEST_SECONDARY_UNIT_SHEETS]
        )
        formula_sheets.extend(
            [
                (LATEST_SECONDARY_SUMMARY_SHEET, 7, 7),
                ("คุณลักษณะ อ่าน สมรรถนะ (9)", 5, 5),
            ]
        )

    for sheet_name, source_row, target_start_row in formula_sheets:
        if sheet_name in wb.sheetnames:
            extend_row_formulas(
                wb[sheet_name], source_row, target_start_row, student_count
            )

    if "ผลการเรียน" in wb.sheetnames:
        extend_row_formulas(
            wb["ผลการเรียน"], 9, 9, student_count, min_col=2, max_col=9
        )


def thai_month_from_label(value):
    text = str(value or "").strip()
    for name, month in THAI_MONTHS.items():
        if name in text:
            return month
    return None


def academic_gregorian_year(assignment, month):
    academic_year = assignment.get("academicYear") or {}
    year = academic_year.get("year")
    try:
        year = int(year)
    except (TypeError, ValueError):
        return None
    if year > 2400:
        year -= 543
    return year + 1 if month <= 3 else year


def parse_attendance_date(value):
    text = str(value or "").strip()
    if not text:
        return None
    match = re.match(r"^(\d{4})-(\d{2})-(\d{2})", text)
    if not match:
        return None
    return f"{match.group(1)}-{match.group(2)}-{match.group(3)}"


def date_key(year, month, day):
    try:
        return date(int(year), int(month), int(day)).isoformat()
    except (TypeError, ValueError):
        return None


def month_for_attendance_column(ws, column):
    label = None
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row == 2 and merged_range.min_col <= column <= merged_range.max_col:
            label = ws.cell(row=2, column=merged_range.min_col).value
            break
    if label is None:
        label = ws.cell(row=2, column=column).value
    return thai_month_from_label(label)


def attendance_date_columns(ws, assignment, previous_sheet_end=None):
    columns = []
    current_month = None
    current_year = None
    previous_day = None
    for col in range(1, ws.max_column + 1):
        raw_day = ws.cell(row=4, column=col).value
        try:
            day = int(raw_day)
        except (TypeError, ValueError):
            continue

        header_month = month_for_attendance_column(ws, col)
        if current_month is None:
            next_date = (
                previous_sheet_end + timedelta(days=1)
                if previous_sheet_end is not None
                else None
            )
            current_month = (
                next_date.month
                if next_date is not None and next_date.day == day
                else header_month
            )
            current_year = (
                next_date.year
                if next_date is not None and next_date.day == day
                else academic_gregorian_year(assignment, current_month)
            )
        elif previous_day is not None and day < previous_day:
            if current_month == 12:
                current_month = 1
                current_year += 1
            else:
                current_month += 1
        elif header_month is not None and previous_day is None:
            current_month = header_month
            current_year = academic_gregorian_year(assignment, current_month)

        if current_month is None or current_year is None:
            previous_day = day
            continue

        key = date_key(current_year, current_month, day)
        if key:
            columns.append((key, col))
        previous_day = day
    return columns


def attendance_sheet_ranges(wb, assignment):
    sheets = []
    previous_sheet_end = None
    for ws in wb.worksheets:
        if not ws.title.startswith("เวลาเรียน"):
            continue
        date_columns = attendance_date_columns(ws, assignment, previous_sheet_end)
        if not date_columns:
            continue
        sheets.append(
            {
                "worksheet": ws,
                "date_columns": date_columns,
                "start": date_columns[0][0],
                "end": date_columns[-1][0],
            }
        )
        previous_sheet_end = date.fromisoformat(date_columns[-1][0])
    return sheets


def write_latest_attendance(wb, payload, visible_students):
    assignment = payload["assignment"]
    attendance_by_student_date = {}
    for item in payload.get("attendance", []):
        key = parse_attendance_date(item.get("date"))
        if not key:
            continue
        attendance_by_student_date[(item.get("studentId"), key)] = item.get("status")

    if not attendance_by_student_date:
        return

    student_rows = {
        student.get("id"): row
        for row, student in enumerate(visible_students, start=6)
    }

    attendance_sheets = attendance_sheet_ranges(wb, assignment)
    assigned_attendance = {sheet["worksheet"].title: {} for sheet in attendance_sheets}

    for (student_id, day_key), status in attendance_by_student_date.items():
        for sheet in attendance_sheets:
            if sheet["start"] <= day_key <= sheet["end"]:
                assigned_attendance[sheet["worksheet"].title][(student_id, day_key)] = status
                break

    for sheet in attendance_sheets:
        ws = sheet["worksheet"]
        date_columns = sheet["date_columns"]
        clear_columns(ws, 6, ws.max_row, [col for _, col in date_columns])
        sheet_attendance = assigned_attendance.get(ws.title, {})
        if not sheet_attendance:
            continue
        for day_key, col in date_columns:
            for student_id, row in student_rows.items():
                status = sheet_attendance.get((student_id, day_key))
                if not status:
                    continue
                write_if_not_merged(ws, row, col, ATTENDANCE_MARKS.get(status, status))


def category_score(score_map, category, student):
    return score_map.get((category.get("id"), student.get("id")))


def get_term_categories(categories, term):
    return [
        category
        for category in categories
        if normalize_primary_term(category) == term
        and not is_primary_fixed_final_category(category)
    ]


def clear_score_entry_columns(ws, columns, start_row, detail_width):
    clear_cols = []
    for col in columns:
        clear_cols.extend(range(col, col + detail_width))
    clear_columns(ws, start_row, ws.max_row, clear_cols)


def write_category_to_column(ws, category, students, score_map, column, header_row, first_student_row):
    write_excel_number(ws.cell(row=header_row, column=column), category.get("maxScore"))
    for row, student in enumerate(students, start=first_student_row):
        write_excel_number(ws.cell(row=row, column=column), category_score(score_map, category, student))


def write_latest_primary_scores(wb, categories, visible_students, score_map):
    summary = wb[LATEST_PRIMARY_SUMMARY_SHEET]
    summary_unit_columns = (
        LATEST_PRIMARY_UNIT_SUMMARY_COLUMNS["midyear"]
        + LATEST_PRIMARY_UNIT_SUMMARY_COLUMNS["endyear"]
    )

    clear_columns(summary, 7, summary.max_row, summary_unit_columns)
    clear_columns(summary, 8, summary.max_row, list(LATEST_PRIMARY_FINAL_SUMMARY_COLUMNS.values()))

    for term, sheet_name in LATEST_PRIMARY_TERM_SHEETS.items():
        ws = wb[sheet_name]
        clear_score_entry_columns(ws, LATEST_PRIMARY_UNIT_SCORE_COLUMNS, start_row=6, detail_width=3)
        term_categories = get_term_categories(categories, term)
        for index, category in enumerate(term_categories[: len(LATEST_PRIMARY_UNIT_SCORE_COLUMNS)]):
            score_col = LATEST_PRIMARY_UNIT_SCORE_COLUMNS[index]
            summary_col = LATEST_PRIMARY_UNIT_SUMMARY_COLUMNS[term][index]
            write_category_to_column(ws, category, visible_students, score_map, score_col, 6, 7)
            write_category_to_column(summary, category, visible_students, score_map, summary_col, 7, 8)

        final_category = get_primary_final_category(categories, term)
        final_summary_col = LATEST_PRIMARY_FINAL_SUMMARY_COLUMNS[term]
        if final_category is not None:
            write_excel_number(summary.cell(row=7, column=final_summary_col), final_category.get("maxScore"))
            for row, student in enumerate(visible_students, start=8):
                write_excel_number(
                    summary.cell(row=row, column=final_summary_col),
                    category_score(score_map, final_category, student),
                )


def write_latest_secondary_scores(wb, categories, visible_students, score_map):
    summary = wb[LATEST_SECONDARY_SUMMARY_SHEET]
    unit_categories = [
        category
        for category in categories
        if not is_primary_fixed_final_category(category)
    ]

    unit_slots = []
    for sheet_name, columns in LATEST_SECONDARY_UNIT_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        clear_score_entry_columns(ws, columns, start_row=5, detail_width=5)
        for col in columns:
            unit_slots.append((ws, col))

    clear_columns(summary, 6, summary.max_row, LATEST_SECONDARY_UNIT_SUMMARY_COLUMNS)
    clear_columns(summary, 7, summary.max_row, list(LATEST_SECONDARY_FINAL_SUMMARY_COLUMNS.values()))

    for index, category in enumerate(unit_categories[: len(unit_slots)]):
        ws, score_col = unit_slots[index]
        summary_col = LATEST_SECONDARY_UNIT_SUMMARY_COLUMNS[index]
        write_category_to_column(ws, category, visible_students, score_map, score_col, 5, 6)
        write_category_to_column(summary, category, visible_students, score_map, summary_col, 6, 7)

    for term, final_col in LATEST_SECONDARY_FINAL_SUMMARY_COLUMNS.items():
        final_category = get_secondary_final_category(categories, term)
        if final_category is not None:
            write_excel_number(summary.cell(row=6, column=final_col), final_category.get("maxScore"))
            for row, student in enumerate(visible_students, start=7):
                write_excel_number(
                    summary.cell(row=row, column=final_col),
                    category_score(score_map, final_category, student),
                )


def fill_latest_academic_print_workbook(wb, payload):
    assignment = payload["assignment"]
    visible_students = [student for student in payload.get("students", []) if is_visible_student(student)]
    categories = sort_primary_categories(payload.get("categories", []))
    score_map = {
        (score.get("categoryId"), score.get("studentId")): score.get("score")
        for score in payload.get("scores", [])
    }

    write_latest_cover(wb, assignment)
    extend_latest_student_formulas(wb, visible_students)
    write_latest_student_lists(wb, visible_students)
    write_latest_score_student_names(wb, visible_students)
    write_latest_attendance(wb, payload, visible_students)
    repair_latest_assessment_formulas(wb, visible_students)
    if has_latest_primary_layout(wb):
        write_latest_primary_scores(wb, categories, visible_students, score_map)
        wb.active = wb.sheetnames.index(LATEST_PRIMARY_SUMMARY_SHEET)
        return True

    if has_latest_secondary_layout(wb):
        write_latest_secondary_scores(wb, categories, visible_students, score_map)
        wb.active = wb.sheetnames.index(LATEST_SECONDARY_SUMMARY_SHEET)
        return True

    return False


def student_visible_name(student):
    return " ".join(
        part.strip()
        for part in [
            str(student.get("prefix", "") or "").strip(),
            str(student.get("firstName", "") or "").strip(),
            str(student.get("lastName", "") or "").strip(),
        ]
        if part.strip()
    ).strip()


def is_visible_student(student):
    return bool(student_visible_name(student))


def write_common_data_sheet(wb, payload):
    ws = ensure_sheet(wb, "ส่งออกข้อมูลจริง")
    ws.delete_rows(1, ws.max_row)

    meta = payload.get("assignment") or payload.get("student") or {}
    title = "ข้อมูลส่งออก"
    if payload.get("mode") == "class":
        title = f"{meta.get('classroomName', '')} {meta.get('subjectCode', '')} {meta.get('subjectName', '')}".strip()
    elif payload.get("mode") == "student":
        title = f"{meta.get('studentCode', '')} {meta.get('firstName', '')} {meta.get('lastName', '')}".strip()

    ws["A1"] = "หัวข้อ"
    ws["B1"] = title
    ws["A2"] = "ชนิด"
    ws["B2"] = payload.get("mode", "")

    row = 4
    if payload.get("mode") == "class":
        tables = [
            ("นักเรียน", payload.get("students", [])),
            ("หมวดคะแนน", payload.get("categories", [])),
            ("คะแนนดิบ", payload.get("scores", [])),
            ("ผลการเรียน", payload.get("gradeResults", [])),
            ("เช็คชื่อ", payload.get("attendance", [])),
        ]
    else:
        tables = [("ผลการเรียนรายบุคคล", payload.get("gradeResults", []))]

    for table_name, table_rows in tables:
        ws.cell(row=row, column=1, value=table_name)
        row += 1
        if not table_rows:
            ws.cell(row=row, column=1, value="ไม่มีข้อมูล")
            row += 2
            continue

        headers = list(table_rows[0].keys())
        for col_index, header in enumerate(headers, start=1):
            ws.cell(row=row, column=col_index, value=header)
        row += 1
        for item in table_rows:
            for col_index, header in enumerate(headers, start=1):
                ws.cell(row=row, column=col_index, value=item.get(header, ""))
            row += 1
        row += 2


def fill_class_workbook(wb, payload, template_name):
    if fill_latest_academic_print_workbook(wb, payload):
        return

    assignment = payload["assignment"]
    visible_students = [student for student in payload.get("students", []) if is_visible_student(student)]
    categories = sort_primary_categories(payload.get("categories", []))
    score_map = {
        (score.get("categoryId"), score.get("studentId")): score.get("score")
        for score in payload.get("scores", [])
    }
    cover = wb[wb.sheetnames[0]]
    if "ปก" in wb.sheetnames:
        cover = wb["ปก"]

    title = f"{assignment.get('classroomName', '')} {assignment.get('subjectCode', '')} {assignment.get('subjectName', '')}".strip()
    if is_secondary_template(template_name, assignment):
        cover["A9"] = f"ชั้นมัธยมศึกษา {assignment.get('classroomName', '')} ปีการศึกษา {assignment.get('academicYear', {}).get('year', '')}"
    elif is_primary_template(template_name, assignment):
        cover["A9"] = f"ชั้นประถมศึกษา {assignment.get('classroomName', '')} ปีการศึกษา {assignment.get('academicYear', {}).get('year', '')}"

    cover["A5"] = title or cover["A5"].value
    if is_primary_template(template_name, assignment) and "ปก" in wb.sheetnames:
        cover["A11"] = (
            f"รายวิชา {assignment.get('subjectName', '')} "
            f"รหัสวิชา {assignment.get('subjectCode', '')} "
            f"หน่วยกิต {assignment.get('subjectCredits', '')} "
            f"{assigned_teacher_label(assignment)} {' / '.join(homeroom_teacher_names(assignment))}"
        ).strip()

    if is_primary_template(template_name, assignment):
        subject_ws = find_subject_sheet(wb, assignment)
        for ws in wb.worksheets:
            if ws.title in {"ปก", "ข้อมูลนร.2-3", "เวลาเรียน4-16", "ส่งออกข้อมูลจริง", "เกรดรวม59-60"}:
                continue
            if ws.title.startswith("คำอธิบาย") or ws.title == "ผลการเรียน":
                continue
            clear_primary_subject_sheet(ws)

        if subject_ws is not None:
            subject_title = f"รายวิชา {assignment.get('subjectName', '')}  ({assignment.get('subjectCode', '')})".strip()
            subject_ws["A3"] = subject_title
            subject_ws["AM3"] = subject_title

            for term in ("midyear", "endyear"):
                term_categories = [
                    category
                    for category in categories
                    if normalize_primary_term(category) == term
                    and not is_primary_fixed_final_category(category)
                ]
                score_columns = get_primary_unit_columns(term)
                for category, col in zip(term_categories, score_columns):
                    write_excel_number(
                        subject_ws.cell(row=6, column=col),
                        category.get("maxScore"),
                    )
                    for row_offset, student in enumerate(visible_students, start=7):
                        write_excel_number(
                            subject_ws.cell(row=row_offset, column=col),
                            score_map.get((category.get("id"), student.get("id"))),
                        )

                final_category = get_primary_final_category(categories, term)
                final_col = PRIMARY_FINAL_COLUMNS[term]
                if final_category is not None:
                    write_excel_number(
                        subject_ws.cell(row=6, column=final_col),
                        final_category.get("maxScore"),
                    )
                    for row_offset, student in enumerate(visible_students, start=7):
                        write_excel_number(
                            subject_ws.cell(row=row_offset, column=final_col),
                            score_map.get((final_category.get("id"), student.get("id"))),
                        )

            subject_index = wb._sheets.index(subject_ws)
            if wb.views:
                wb.views[0].activeTab = subject_index
                wb.views[0].firstSheet = subject_index
            wb.active = subject_index
            subject_ws.sheet_view.tabSelected = True

    if "ข้อมูลนร.2-3" in wb.sheetnames:
        ws = wb["ข้อมูลนร.2-3"]
        clear_rect(ws, 4, ws.max_row, 1, 16)
        for idx, student in enumerate(visible_students, start=4):
            ws.cell(row=idx, column=1, value=student.get("studentNumber") or "")
            ws.cell(row=idx, column=2, value=student.get("studentCode") or "")
            ws.cell(row=idx, column=3, value=student.get("nationalId") or "")
            ws.cell(row=idx, column=4, value=student_visible_name(student))
            ws.cell(row=idx, column=5, value=student.get("birthDate") or "")

    if "เวลาเรียน4-16" in wb.sheetnames:
        ws = wb["เวลาเรียน4-16"]
        start_row = 6 + len(visible_students)
        if start_row <= ws.max_row:
            clear_rect(ws, start_row, ws.max_row, 1, ws.max_column)

    if "รายงานนักเรียนรายบุคคล" in wb.sheetnames and payload.get("firstStudent"):
        student = payload["firstStudent"]
        ws = wb["รายงานนักเรียนรายบุคคล"]
        ws["D6"] = student_visible_name(student)
        ws["G6"] = student.get("studentCode", "")
        ws["I6"] = student.get("studentNumber", "")
        ws["N6"] = student.get("studentCode", "")
        ws["O6"] = student_visible_name(student)
        clear_rect(ws, 8, 200, 2, 10)
        for idx, result in enumerate(payload.get("gradeResults", []), start=8):
            # Fill the first visible column set with the current assignment data.
            ws.cell(row=idx, column=2, value="พื้นฐาน")
            ws.cell(row=idx, column=3, value=assignment.get("subjectCode", ""))
            ws.cell(row=idx, column=4, value=assignment.get("subjectName", ""))
            ws.cell(row=idx, column=5, value=result.get("totalHours", ""))
            ws.cell(row=idx, column=6, value="100")
            ws.cell(row=idx, column=7, value=result.get("classAverage", ""))
            ws.cell(row=idx, column=8, value=result.get("totalScore", ""))
            ws.cell(row=idx, column=9, value=result.get("grade", ""))
            ws.cell(row=idx, column=10, value=result.get("result", ""))

def fill_student_workbook(wb, payload, template_name):
    student = payload["student"]
    if "รายงานนักเรียนรายบุคคล" in wb.sheetnames:
        ws = wb["รายงานนักเรียนรายบุคคล"]
        ws["D6"] = f"{student.get('prefix', '')}{student.get('firstName', '')} {student.get('lastName', '')}".strip()
        ws["G6"] = student.get("studentCode", "")
        ws["I6"] = student.get("id", "")
        ws["N4"] = f"เริ่มพิมพ์จากเลขที่ {student.get('studentCode', '')}"
        ws["N6"] = student.get("studentCode", "")
        ws["O6"] = f"{student.get('prefix', '')}{student.get('firstName', '')} {student.get('lastName', '')}".strip()
        clear_rect(ws, 8, 200, 2, 10)
        for idx, result in enumerate(payload.get("gradeResults", []), start=8):
            ws.cell(row=idx, column=2, value="พื้นฐาน")
            ws.cell(row=idx, column=3, value=result.get("subjectCode", ""))
            ws.cell(row=idx, column=4, value=result.get("subjectName", ""))
            ws.cell(row=idx, column=5, value=result.get("totalHours", ""))
            ws.cell(row=idx, column=6, value="100")
            ws.cell(row=idx, column=7, value=result.get("classAverage", ""))
            ws.cell(row=idx, column=8, value=result.get("totalScore", ""))
            ws.cell(row=idx, column=9, value=result.get("grade", ""))
            ws.cell(row=idx, column=10, value=result.get("result", ""))


def copy_styles(source_ws, target_ws):
    for row in source_ws.iter_rows():
        for cell in row:
            target_cell = target_ws[cell.coordinate]
            if cell.has_style:
                target_cell._style = copy(cell._style)
            if cell.number_format:
                target_cell.number_format = cell.number_format
            if cell.font:
                target_cell.font = copy(cell.font)
            if cell.fill:
                target_cell.fill = copy(cell.fill)
            if cell.border:
                target_cell.border = copy(cell.border)
            if cell.alignment:
                target_cell.alignment = copy(cell.alignment)
            if cell.protection:
                target_cell.protection = copy(cell.protection)


def main():
    if len(sys.argv) < 3:
        print("usage: excel_exporter.py <template-path> <output-path>", file=sys.stderr)
        return 1

    template_path = sys.argv[1]
    output_path = sys.argv[2]
    payload = json.loads(sys.stdin.read() or "{}")

    keep_vba = template_path.lower().endswith(".xlsm")
    wb = load_workbook(template_path, keep_vba=keep_vba)

    write_common_data_sheet(wb, payload)

    if payload.get("mode") == "class":
        fill_class_workbook(wb, payload, os.path.basename(template_path))
    else:
        fill_student_workbook(wb, payload, os.path.basename(template_path))

    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    wb.save(output_path)
    return 0


if __name__ == "__main__":
    sys.exit(main())
