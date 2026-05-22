#!/usr/bin/env python3
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


HEADERS = [
    "เลขที่",
    "รหัสนักเรียน",
    "คำนำหน้า",
    "ชื่อ",
    "นามสกุล",
    "เลขประชาชน",
    "วันเกิด",
    "เพศ",
    "สถานะ",
]

EXAMPLES = [
    [1, "1234", "เด็กชาย", "สมชาย", "รักเรียน", "", "2016-05-21", "ชาย", "ปกติ"],
    [2, "1235", "เด็กหญิง", "สมหญิง", "ตั้งใจ", "", "2016-08-10", "หญิง", "ปกติ"],
]


def main():
    if len(sys.argv) < 2:
        print("usage: student_import_template.py <output-path>", file=sys.stderr)
        return 1

    output_path = sys.argv[1]
    wb = Workbook()
    ws = wb.active
    ws.title = "รายชื่อนักเรียน"

    ws["A1"] = "เทมเพลตนำเข้ารายชื่อนักเรียน"
    ws["A1"].font = Font(bold=True, size=16, color="1E3A8A")
    ws.merge_cells("A1:I1")

    ws["A2"] = "กรอกข้อมูลตั้งแต่แถวที่ 5 เป็นต้นไป ช่องที่จำเป็นคือ รหัสนักเรียน, ชื่อ, นามสกุล"
    ws["A2"].font = Font(color="475569")
    ws.merge_cells("A2:I2")

    header_fill = PatternFill("solid", fgColor="DBEAFE")
    border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    header_row = 4
    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True, color="0F172A")
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_index, row_values in enumerate(EXAMPLES, start=5):
        for col, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_index, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    # Add blank rows with the same border so the template feels ready to fill.
    for row_index in range(7, 57):
        for col in range(1, len(HEADERS) + 1):
            ws.cell(row=row_index, column=col).border = border

    widths = [8, 18, 14, 18, 22, 20, 14, 10, 14]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = "A4:I56"

    prefix_validation = DataValidation(type="list", formula1='"เด็กชาย,เด็กหญิง,นาย,นางสาว"', allow_blank=True)
    gender_validation = DataValidation(type="list", formula1='"ชาย,หญิง"', allow_blank=True)
    status_validation = DataValidation(type="list", formula1='"ปกติ,ย้าย,จบ,ลาออก"', allow_blank=True)
    ws.add_data_validation(prefix_validation)
    ws.add_data_validation(gender_validation)
    ws.add_data_validation(status_validation)
    prefix_validation.add("C5:C200")
    gender_validation.add("H5:H200")
    status_validation.add("I5:I200")

    note = wb.create_sheet("คำอธิบาย")
    note["A1"] = "คำอธิบายคอลัมน์"
    note["A1"].font = Font(bold=True, size=14)
    rows = [
        ("เลขที่", "เลขที่ในห้อง เช่น 1, 2, 3"),
        ("รหัสนักเรียน", "จำเป็น และต้องไม่ซ้ำ เช่น 1234"),
        ("คำนำหน้า", "เด็กชาย, เด็กหญิง, นาย, นางสาว"),
        ("ชื่อ", "จำเป็น"),
        ("นามสกุล", "จำเป็น"),
        ("เลขประชาชน", "ใส่หรือเว้นว่างได้"),
        ("วันเกิด", "แนะนำรูปแบบ yyyy-mm-dd เช่น 2016-05-21"),
        ("เพศ", "ชาย หรือ หญิง"),
        ("สถานะ", "ปกติ, ย้าย, จบ, ลาออก ถ้าเว้นว่างจะถือว่า ปกติ"),
    ]
    for row_index, (name, desc) in enumerate(rows, start=3):
        note.cell(row=row_index, column=1, value=name).font = Font(bold=True)
        note.cell(row=row_index, column=2, value=desc)
    note.column_dimensions["A"].width = 18
    note.column_dimensions["B"].width = 60

    wb.save(output_path)
    return 0


if __name__ == "__main__":
    sys.exit(main())
